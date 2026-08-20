"""Estilos profesionales para los documentos generados."""
from __future__ import annotations
from contextlib import contextmanager
from contextvars import ContextVar
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.styles.numbers import FORMAT_DATE_DATETIME

# Paleta base — colores estructurales/neutros, no configurables desde Configuración.
C_DARK_BG   = "091F2C"   # Azul oscuro principal
C_HEADER_BG = "091F2C"   # Cabeceras principales
C_ACCENT    = "4F758B"   # Azul medio
C_LIGHT_BG  = "DAE4EA"   # Azul claro tenue (filas alternas — A6BBC8 aclarado)
C_WHITE     = "FFFFFF"
C_BLACK     = "000000"
C_GRAY_LIGHT = "A6BBC8"  # Azul claro
C_GRAY_MED  = "7A99AC"   # Azul gris (tagline)
C_TEXT_DARK = "091F2C"   # Texto sobre fondos claros

# ─── Colores de informe configurables (Configuración → Edición de informes) ────
# STYLE_TITLE/STYLE_HEADER/STYLE_SUBHEADER leen estos valores en el momento de
# construir el dict (no al importar el módulo), vía ContextVar — así cada
# generación de informe puede fijar sus propios colores (report_colors())
# sin interferir con otras peticiones concurrentes ni tener que pasar el color
# como parámetro explícito por las ~15 funciones _sheet_* de excel.py.
DEFAULT_HEADER_COLOR = "15121F"
DEFAULT_ACCENT_COLOR = "7C3AED"
DEFAULT_SEPARATOR_COLOR = "8B5CF6"

_header_ctx: ContextVar[str] = ContextVar("report_header_color", default=DEFAULT_HEADER_COLOR)
_accent_ctx: ContextVar[str] = ContextVar("report_accent_color", default=DEFAULT_ACCENT_COLOR)
_separator_ctx: ContextVar[str] = ContextVar("report_separator_color", default=DEFAULT_SEPARATOR_COLOR)


def get_header_color() -> str:
    return _header_ctx.get()


def get_accent_color() -> str:
    return _accent_ctx.get()


def get_separator_color() -> str:
    return _separator_ctx.get()


@contextmanager
def report_colors(header: str, accent: str, separator: str):
    """Fija los 3 colores de informe para la duración de la generación actual."""
    h_tok = _header_ctx.set(header)
    a_tok = _accent_ctx.set(accent)
    s_tok = _separator_ctx.set(separator)
    try:
        yield
    finally:
        _header_ctx.reset(h_tok)
        _accent_ctx.reset(a_tok)
        _separator_ctx.reset(s_tok)

# Colores de severidad — paleta LÃBERIT por orden de urgencia
C_CRITICAL = "C10016"   # Rojo corporativo
C_HIGH     = "091F2C"   # Azul oscuro (alta urgencia)
C_MEDIUM   = "4F758B"   # Azul medio
C_LOW      = "7A99AC"   # Azul gris
C_INFO     = "A6BBC8"   # Azul claro (requiere texto oscuro)

# Colores de estado — paleta LÃBERIT
C_OK      = "4F758B"   # Azul medio (estable)
C_WARNING = "B4B5DF"   # Lavanda (alerta suave, requiere texto oscuro)
C_ERROR   = "C10016"   # Rojo corporativo


def _side(style="thin", color=C_GRAY_MED):
    return Side(style=style, color=color)


def _border(all_sides="thin"):
    s = _side(all_sides)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=10, color=C_TEXT_DARK, italic=False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic)


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# Estilos predefinidos — TITLE/HEADER/SUBHEADER son funciones (no dicts) porque
# su color de fondo depende del informe en curso (ver report_colors() arriba).
def STYLE_TITLE() -> dict:
    return {
        "font": _font(bold=True, size=20, color=C_WHITE),
        "fill": _fill(get_header_color()),
        "alignment": _align("center", "center"),
        # Línea separadora bajo la cabecera (equivalente Excel al HRFlowable del PDF).
        "border": Border(bottom=_side("medium", get_separator_color())),
    }


STYLE_SUBTITLE = {
    "font": _font(bold=True, size=12, color=C_ACCENT),
    "fill": _fill(C_WHITE),
    "alignment": _align("left", "center"),
}


def STYLE_HEADER() -> dict:
    return {
        "font": _font(bold=True, size=10, color=C_WHITE),
        "fill": _fill(get_header_color()),
        "alignment": _align("center", "center"),
        "border": _border(),
    }


def STYLE_SUBHEADER() -> dict:
    return {
        "font": _font(bold=True, size=10, color=C_WHITE),
        "fill": _fill(get_accent_color()),
        "alignment": _align("center", "center"),
        "border": _border(),
    }

STYLE_ROW_EVEN = {
    "font": _font(size=9),
    "fill": _fill(C_WHITE),
    "alignment": _align(wrap=True),
    "border": _border(),
}

STYLE_ROW_ODD = {
    "font": _font(size=9),
    "fill": _fill(C_LIGHT_BG),
    "alignment": _align(wrap=True),
    "border": _border(),
}

STYLE_LABEL = {
    "font": _font(bold=True, size=9, color=C_TEXT_DARK),
    "fill": _fill(C_GRAY_LIGHT),
    "alignment": _align(),
    "border": _border(),
}

STYLE_VALUE = {
    "font": _font(size=9),
    "fill": _fill(C_WHITE),
    "alignment": _align(wrap=True),
    "border": _border(),
}

# Severidad — texto oscuro en fondos claros (informational = A6BBC8)
SEVERITY_STYLES = {
    "critical":     {"font": _font(bold=True, size=9, color=C_WHITE),      "fill": _fill(C_CRITICAL)},
    "high":         {"font": _font(bold=True, size=9, color=C_WHITE),      "fill": _fill(C_HIGH)},
    "medium":       {"font": _font(bold=True, size=9, color=C_WHITE),      "fill": _fill(C_MEDIUM)},
    "low":          {"font": _font(bold=True, size=9, color=C_WHITE),      "fill": _fill(C_LOW)},
    "informational":{"font": _font(size=9,           color=C_TEXT_DARK),   "fill": _fill(C_INFO)},
}

# Estado — texto oscuro en lavanda (warning = B4B5DF, contraste bajo con blanco)
STATUS_STYLES = {
    "ok":      {"font": _font(bold=True, size=9, color=C_WHITE),      "fill": _fill(C_OK)},
    "success": {"font": _font(bold=True, size=9, color=C_WHITE),      "fill": _fill(C_OK)},
    "warning": {"font": _font(bold=True, size=9, color=C_TEXT_DARK),  "fill": _fill(C_WARNING)},
    "failed":  {"font": _font(bold=True, size=9, color=C_WHITE),      "fill": _fill(C_ERROR)},
    "error":   {"font": _font(bold=True, size=9, color=C_WHITE),      "fill": _fill(C_ERROR)},
}


def apply_style(cell, style: dict) -> None:
    if "font" in style:
        cell.font = style["font"]
    if "fill" in style:
        cell.fill = style["fill"]
    if "alignment" in style:
        cell.alignment = style["alignment"]
    if "border" in style:
        cell.border = style["border"]


SEVERITY_LABELS = {
    "critical": "CRÍTICO",
    "high": "ALTO",
    "medium": "MEDIO",
    "low": "BAJO",
    "informational": "INFORMATIVO",
}

DEVICE_TYPE_LABELS = {
    "windows_server": "Windows Server",
    "windows_workstation": "Windows PC",
    "linux": "Linux",
    "esxi": "VMware ESXi",
    "vcenter": "VMware vCenter",
    "fortigate": "FortiGate",
    "nas": "NAS",
    "switch": "Switch",
    "router": "Router",
    "ilo": "HP iLO",
    "idrac": "Dell iDRAC",
    "printer": "Impresora",
    "veeam": "Veeam",
    "unknown": "Desconocido",
}
