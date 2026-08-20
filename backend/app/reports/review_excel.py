"""Generación de informe Excel para revisiones manuales."""
from __future__ import annotations
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage


def _add_xl_logo(ws, logo_path: Path, anchor: str, max_w: int, max_h: int) -> None:
    """Inserta una imagen escalada proporcionalmente en la hoja Excel."""
    if not logo_path or not logo_path.exists():
        return
    try:
        img = XLImage(str(logo_path))
        nat_w, nat_h = img.width, img.height
        if nat_w and nat_h:
            ratio = min(max_w / nat_w, max_h / nat_h)
            img.width  = max(1, int(nat_w * ratio))
            img.height = max(1, int(nat_h * ratio))
        else:
            img.width, img.height = max_w, max_h
        ws.add_image(img, anchor)
    except Exception:
        pass

_INVALID_SHEET_CHARS = re.compile(r'[\\/?*\[\]:]')


def _safe_sheet_title(label: str) -> str:
    """Sanitize sheet title: remove forbidden chars, max 31 chars."""
    safe = _INVALID_SHEET_CHARS.sub('', label).strip()
    return safe[:31] or "Hoja"

from .styles import (
    apply_style, STYLE_TITLE, STYLE_HEADER, STYLE_SUBHEADER,
    STYLE_LABEL, STYLE_VALUE, STYLE_ROW_EVEN,
    _fill, _font, _border, _align,
    C_OK, C_WARNING, C_ERROR, C_WHITE, C_TEXT_DARK, C_CRITICAL,
    report_colors, get_accent_color,
    DEFAULT_HEADER_COLOR, DEFAULT_ACCENT_COLOR, DEFAULT_SEPARATOR_COLOR,
)
from .review_items import REVIEW_CATEGORIES
from .report_branding import get_report_logo_path, get_report_colors, REPORT_TAGLINE
from ..services.review_checklist import get_effective_items

_STATUS_SYMBOL = {"ok": "✓", "warning": "!", "critical": "✕", "": ""}

# Fills para la tabla de resumen (estado general del dispositivo — mantiene fondo)
_STATUS_FILL   = {"ok": _fill(C_OK), "warning": _fill(C_WARNING), "critical": _fill(C_ERROR), "": _fill("DAE4EA")}
_STATUS_FONT   = {
    "ok":       _font(bold=True, size=10, color=C_WHITE),
    "warning":  _font(bold=True, size=10, color=C_TEXT_DARK),
    "critical": _font(bold=True, size=10, color=C_WHITE),
    "":         _font(size=9, color="A6BBC8"),
}

# Checks en tablas de ítems: sin fondo, solo el color del tick
_CHECK_FONT = {
    "ok":       _font(bold=True, size=11, color="091F2C"),   # azul oscuro
    "warning":  _font(bold=True, size=11, color="7A99AC"),   # azul gris
    "critical": _font(bold=True, size=11, color=C_CRITICAL), # rojo
}
_STATUS_LABEL  = {"ok": "OK", "warning": "Warning", "critical": "Critical", "": "—"}

_ODD_ROW = {
    "font": _font(size=9),
    "fill": _fill("DAE4EA"),   # Azul claro tenue (filas alternas)
    "alignment": _align(wrap=True),
    "border": _border(),
}


def _cell(ws, col: int, row: int, value="", style: dict | None = None):
    c = ws.cell(row=row, column=col, value=value)
    if style:
        apply_style(c, style)
    return c


def _write_info_header(ws, client_name: str, technician_name: str, review_date: str, category_label: str) -> None:
    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1, value=f"REVISIÓN MANUAL — {category_label.upper()}")
    apply_style(c, STYLE_TITLE())
    ws.row_dimensions[1].height = 28

    _cell(ws, 1, 2, "Cliente:", STYLE_LABEL)
    ws.merge_cells("B2:D2")
    _cell(ws, 2, 2, client_name, STYLE_VALUE)
    _cell(ws, 5, 2, "FECHA:", STYLE_LABEL)
    _cell(ws, 6, 2, review_date, STYLE_VALUE)

    _cell(ws, 1, 3, "Técnico:", STYLE_LABEL)
    ws.merge_cells("B3:D3")
    _cell(ws, 2, 3, technician_name, STYLE_VALUE)


def _write_category_sheet(
    wb: Workbook,
    category_key: str,
    category_label: str,
    devices_data: list[dict],
    client_name: str,
    technician_name: str,
    review_date: str,
    removed_items: dict | None = None,
    custom_items: dict | None = None,
) -> None:
    ws = wb.create_sheet(title=_safe_sheet_title(category_label))

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 45

    _write_info_header(ws, client_name, technician_name, review_date, category_label)

    row = 5
    for device_entry in devices_data:
        device = device_entry["device"]
        cat_result = device_entry.get("category_result", {})
        items_results: dict = cat_result.get("items", {})
        observations: str = cat_result.get("observations", "")

        device_name = device.get("display_name") or device.get("hostname") or device.get("ip_address", "")
        ip = device.get("ip_address", "")
        device_type = device.get("device_type", "")

        # Device header
        ws.merge_cells(f"A{row}:F{row}")
        c = ws.cell(row=row, column=1, value=f"▸  {device_name}")
        apply_style(c, {
            "font": _font(bold=True, size=11, color=C_WHITE),
            "fill": _fill(get_accent_color()),   # banner de dispositivo — acento de informe
            "alignment": _align("left", "center"),
            "border": _border(),
        })
        ws.row_dimensions[row].height = 20
        row += 1

        _cell(ws, 1, row, "IP:", STYLE_LABEL)
        ws.merge_cells(f"B{row}:F{row}")
        _cell(ws, 2, row, ip, STYLE_VALUE)
        row += 1

        # Column headers
        for col, h in enumerate(["Ítem de revisión", "OK", "Warning", "Critical", "", ""], 1):
            c = ws.cell(row=row, column=col, value=h)
            apply_style(c, STYLE_HEADER())
        row += 1

        cat_items = get_effective_items(category_key, device_type, removed_items, custom_items)
        general_status = items_results.get("general", "")
        if not cat_items:
            if general_status:
                _cell(ws, 1, row, "Estado general", STYLE_ROW_EVEN)
                for col, s in enumerate(["ok", "warning", "critical"], 2):
                    c = ws.cell(row=row, column=col)
                    if general_status == s:
                        c.value = _STATUS_SYMBOL[s]
                        c.fill = _fill("FFFFFF")
                        c.font = _CHECK_FONT[s]
                    else:
                        c.value = ""
                        c.fill = _fill("FFFFFF")
                        c.font = _font(size=9)
                    c.alignment = _align("center")
                    c.border = _border()
                for col in [5, 6]:
                    _cell(ws, col, row, "", STYLE_ROW_EVEN)
                row += 1
            else:
                ws.merge_cells(f"A{row}:F{row}")
                c = ws.cell(row=row, column=1, value="(Sin ítems definidos para este tipo de dispositivo)")
                apply_style(c, {**STYLE_VALUE, "alignment": _align("center")})
                row += 1
        else:
            for i, item in enumerate(cat_items):
                status = items_results.get(item["key"], "")
                row_style = STYLE_ROW_EVEN if i % 2 == 0 else _ODD_ROW
                label = item["label"] + (" (personalizado)" if item.get("is_custom") else "")

                _cell(ws, 1, row, label, row_style)

                for col, s in enumerate(["ok", "warning", "critical"], 2):
                    c = ws.cell(row=row, column=col)
                    if status == s:
                        c.value = _STATUS_SYMBOL[s]
                        c.fill = _fill("FFFFFF")
                        c.font = _CHECK_FONT[s]
                    else:
                        c.value = ""
                        c.fill = _fill("FFFFFF")
                        c.font = _font(size=9)
                    c.alignment = _align("center")
                    c.border = _border()

                for col in [5, 6]:
                    _cell(ws, col, row, "", row_style)

                row += 1

        if observations:
            _cell(ws, 1, row, "Observaciones:", STYLE_LABEL)
            ws.merge_cells(f"B{row}:F{row}")
            c = ws.cell(row=row, column=2, value=observations)
            apply_style(c, {**STYLE_VALUE, "alignment": _align(wrap=True)})
            ws.row_dimensions[row].height = max(20, min(80, len(observations) // 4 + 15))
            row += 1

        row += 1  # blank row between devices


def generate_review_excel(review_export_data: dict, output_path: Path) -> None:
    colors = review_export_data.get("report_colors") or {}
    with report_colors(
        colors.get("header", DEFAULT_HEADER_COLOR),
        colors.get("accent", DEFAULT_ACCENT_COLOR),
        colors.get("separator", DEFAULT_SEPARATOR_COLOR),
    ):
        _generate_review_excel(review_export_data, output_path)


def _generate_review_excel(review_export_data: dict, output_path: Path) -> None:
    wb = Workbook()
    active = wb.active
    if active is not None:
        wb.remove(active)

    client_name = review_export_data.get("client_name", "")
    technician_name = review_export_data.get("technician_name", "")
    review_date = review_export_data.get("review_date", "")
    categories: list[str] = review_export_data.get("categories", [])
    devices: list[dict] = review_export_data.get("devices", [])
    results: dict = review_export_data.get("results", {})
    device_categories: dict | None = review_export_data.get("device_categories")
    removed_items: dict = review_export_data.get("removed_items") or {}
    custom_items: dict = review_export_data.get("custom_items") or {}
    category_labels: dict = review_export_data.get("category_labels") or {}
    logo_path: str | None = review_export_data.get("logo_path")
    corp_logo = get_report_logo_path()

    def _dev_cats(dev_id: str) -> list[str]:
        if device_categories and dev_id in device_categories:
            return device_categories[dev_id]
        return categories

    # ── Hoja de resumen ────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet(title="Resumen")
    ws_sum.column_dimensions["A"].width = 28
    ws_sum.column_dimensions["B"].width = 28
    ws_sum.column_dimensions["C"].width = 15
    ws_sum.column_dimensions["D"].width = 35
    ws_sum.column_dimensions["E"].width = 18

    # ── Cabecera de branding (filas 1-2): logo de informe izquierda, cliente derecha ──
    ws_sum.row_dimensions[1].height = 42   # espacio para los logos
    ws_sum.row_dimensions[2].height = 8    # separador visual

    if corp_logo:
        _add_xl_logo(ws_sum, corp_logo, anchor="A1", max_w=150, max_h=52)
    if logo_path:
        _add_xl_logo(ws_sum, Path(logo_path), anchor="D1", max_w=120, max_h=52)

    # ── Título (fila 3) ───────────────────────────────────────────────────────
    ws_sum.merge_cells("A3:D3")
    c = ws_sum.cell(row=3, column=1, value="REVISIÓN MANUAL — RESUMEN")
    apply_style(c, STYLE_TITLE())
    ws_sum.row_dimensions[3].height = 28

    # ── Info del informe (filas 4-7) ──────────────────────────────────────────
    info_rows = [
        ("Cliente:", client_name),
        ("Técnico:", technician_name),
        ("Fecha:", review_date),
        ("Categorías:", ", ".join(ck.upper() for ck in categories)),
    ]
    for i, (label, value) in enumerate(info_rows, 4):   # filas 4-7
        _cell(ws_sum, 1, i, label, STYLE_LABEL)
        ws_sum.merge_cells(f"B{i}:D{i}")
        _cell(ws_sum, 2, i, value, STYLE_VALUE)

    row = 9   # era 7; desplazado +2 por la cabecera de branding
    ws_sum.merge_cells(f"A{row}:D{row}")
    c = ws_sum.cell(row=row, column=1, value="RESULTADO POR DISPOSITIVO")
    apply_style(c, STYLE_SUBHEADER())
    row += 1

    for col, h in enumerate(["Dispositivo", "IP", "Estado General", "Observaciones"], 1):
        apply_style(ws_sum.cell(row=row, column=col, value=h), STYLE_HEADER())
    row += 1

    priority = {"critical": 3, "warning": 2, "ok": 1, "": 0}

    for i, device in enumerate(devices):
        dev_id = str(device["id"])
        dev_results = results.get(dev_id, {})
        dev_cats = _dev_cats(dev_id)

        worst = ""
        for cat_key in dev_cats:
            for s in dev_results.get(cat_key, {}).get("items", {}).values():
                if priority.get(s, 0) > priority.get(worst, 0):
                    worst = s

        obs_parts = [
            dev_results[cat].get("observations", "")
            for cat in dev_cats
            if dev_results.get(cat, {}).get("observations")
        ]

        device_name = device.get("display_name") or device.get("hostname") or device.get("ip_address", "")
        row_style = STYLE_ROW_EVEN if i % 2 == 0 else _ODD_ROW

        _cell(ws_sum, 1, row, device_name, row_style)
        _cell(ws_sum, 2, row, device["ip_address"], row_style)

        status_cell = ws_sum.cell(row=row, column=3, value=_STATUS_LABEL.get(worst, "—"))
        status_cell.fill = _STATUS_FILL.get(worst, _fill("DAE4EA"))
        status_cell.font = _STATUS_FONT.get(worst, _font(size=9))
        status_cell.alignment = _align("center")
        status_cell.border = _border()

        _cell(ws_sum, 4, row, " | ".join(obs_parts), {**row_style, "alignment": _align(wrap=True)})
        row += 1

    # ── Hojas por categoría ────────────────────────────────────────────────────
    cat_labels = {ck["key"]: ck["label"] for ck in REVIEW_CATEGORIES}
    cat_labels.update(category_labels)

    for cat_key in categories:
        cat_label = cat_labels.get(cat_key, cat_key.title())
        devices_data = []
        for device in devices:
            dev_id = str(device["id"])
            # Only include devices assigned to this category
            if cat_key not in _dev_cats(dev_id):
                continue
            cat_result = results.get(dev_id, {}).get(cat_key, {"items": {}, "observations": ""})
            devices_data.append({"device": device, "category_result": cat_result})

        if not devices_data:
            continue

        _write_category_sheet(
            wb, cat_key, cat_label, devices_data, client_name, technician_name, review_date,
            removed_items=removed_items, custom_items=custom_items,
        )

    # Pie de página en todas las hojas (visible en impresión/vista previa)
    for ws in wb.worksheets:
        ws.oddFooter.left.text  = REPORT_TAGLINE
        ws.oddFooter.right.text = "Pág. &P / &N"
        ws.page_margins.bottom = 1.2   # pulgadas — espacio para el pie

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
