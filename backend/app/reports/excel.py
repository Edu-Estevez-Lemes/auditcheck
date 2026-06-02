"""
Generador de informes Excel profesionales para ÃUDITCHECK.
14 hojas: Resumen, Inventario, Red y Puertos, Hardware, Backups,
VMware, FortiGate, SNMP, Windows, NAS, iLO/iDRAC, Hallazgos,
Comparativa, Recomendaciones + hoja oculta de metadatos.
"""
from __future__ import annotations
import io
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# Caracteres de control que XML/openpyxl no admite en celdas
_ILLEGAL_CHARS_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f￾￿]')

from .styles import (
    apply_style, STYLE_TITLE, STYLE_HEADER, STYLE_SUBHEADER,
    STYLE_ROW_EVEN, STYLE_ROW_ODD, STYLE_LABEL, STYLE_VALUE,
    SEVERITY_STYLES, SEVERITY_LABELS, DEVICE_TYPE_LABELS,
    STATUS_STYLES, C_HEADER_BG, C_ACCENT, C_WHITE,
)
from ..config import settings
from sqlalchemy.orm import Session
from ..models.audit import Audit
from ..models.device import Device, Port
from ..models.finding import Finding
from ..models.client import Client
from ..models.user import User


def _set_col_width(ws: Worksheet, col: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(col)].width = width


def _clean(val: object) -> object:
    """Elimina caracteres de control ilegales en valores de texto."""
    if isinstance(val, str):
        return _ILLEGAL_CHARS_RE.sub('', val)
    return val


def _write_row(ws: Worksheet, row: int, values: list, style: dict) -> None:
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=_clean(val))
        apply_style(cell, style)


def _add_logo(ws: Worksheet, logo_path: Path | None, anchor: str = "A1") -> None:
    if logo_path and logo_path.exists():
        try:
            img = XLImage(str(logo_path))
            img.width = 120
            img.height = 50
            ws.add_image(img, anchor)
        except Exception:
            pass


def generate_excel_report(
    db: Session,
    audit_id: int,
    output_path: Path | None = None,
    comparison_data: dict | None = None,
) -> bytes:
    audit = db.query(Audit).filter(Audit.id == audit_id).first()
    if not audit:
        raise ValueError(f"Auditoría {audit_id} no encontrada")

    client = db.query(Client).filter(Client.id == audit.client_id).first()
    technician = db.query(User).filter(User.id == audit.technician_id).first() if audit.technician_id else None

    devices = db.query(Device).filter(Device.audit_id == audit_id).all()
    findings = db.query(Finding).filter(Finding.audit_id == audit_id).all()

    client_logo = Path(client.logo_path) if client and client.logo_path else None
    corp_logo = settings.BRANDING_DIR / "logo.png"

    wb = Workbook()
    wb.remove(wb.active)  # Eliminar hoja por defecto

    _sheet_resumen(wb, audit, client, technician, devices, findings, client_logo, corp_logo)
    _sheet_inventario(wb, devices)
    _sheet_red_puertos(wb, db, devices, audit_id)
    _sheet_hallazgos(wb, findings, devices)
    _sheet_recomendaciones(wb, findings)
    if comparison_data:
        _sheet_comparativa(wb, comparison_data)
    else:
        _sheet_comparativa_empty(wb)
    _sheet_hardware(wb)
    _sheet_backups(wb)
    _sheet_vmware(wb)
    _sheet_fortigate(wb)
    _sheet_snmp(wb)
    _sheet_windows(wb)
    _sheet_nas(wb)
    _sheet_ilo_idrac(wb)
    _sheet_metadata(wb, audit, client, technician)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    data = buf.getvalue()

    if output_path:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(data)

    return data


def _write_info_pair(ws, row: int, label: str, value: str, col_start: int = 1, col_end: int = 4):
    ws.cell(row=row, column=col_start, value=label)
    apply_style(ws.cell(row=row, column=col_start), STYLE_LABEL)
    ws.merge_cells(f"{get_column_letter(col_start + 1)}{row}:{get_column_letter(col_end)}{row}")
    ws.cell(row=row, column=col_start + 1, value=_clean(value))
    apply_style(ws.cell(row=row, column=col_start + 1), STYLE_VALUE)


def _sheet_resumen(wb, audit, client, technician, devices, findings, client_logo, corp_logo):
    ws = wb.create_sheet("Resumen General")
    ws.sheet_view.showGridLines = False

    # Logos — corporativo izquierda, cliente derecha
    _add_logo(ws, corp_logo, "A1")
    if client_logo:
        _add_logo(ws, client_logo, "G1")

    # Título
    ws.row_dimensions[4].height = 40
    ws.merge_cells("A4:H4")
    cell = ws["A4"]
    cell.value = "INFORME DE AUDITORÍA TÉCNICA"
    apply_style(cell, STYLE_TITLE)

    row = 6

    # ── Sección: Datos del cliente ──────────────────────────────────────────────
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="DATOS DEL CLIENTE")
    apply_style(ws.cell(row=row, column=1), STYLE_SUBHEADER)
    row += 1

    client_pairs = [
        ("Empresa / Cliente:", client.name if client else ""),
        ("CIF / NIF:", client.cif_nif or ""),
        ("Dirección:", client.address or ""),
        ("Persona de contacto:", client.contact_person or ""),
        ("Teléfono:", client.phone or ""),
        ("Email:", client.email or ""),
    ]
    if client and client.observations:
        client_pairs.append(("Observaciones:", client.observations))

    for label, value in client_pairs:
        _write_info_pair(ws, row, label, value, col_start=1, col_end=5)
        row += 1

    row += 1

    # ── Sección: Datos de la revisión ───────────────────────────────────────────
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="DATOS DE LA REVISIÓN")
    apply_style(ws.cell(row=row, column=1), STYLE_SUBHEADER)
    row += 1

    audit_pairs = [
        ("Nombre revisión:", audit.name),
        ("Técnico responsable:", technician.full_name if technician else ""),
        ("Rangos auditados:", audit.scanned_ranges or ""),
        ("Fecha inicio:", audit.started_at.strftime("%d/%m/%Y %H:%M") if audit.started_at else ""),
        ("Fecha finalización:", audit.completed_at.strftime("%d/%m/%Y %H:%M") if audit.completed_at else ""),
        ("Versión AUDITCHECK:", audit.app_version or settings.APP_VERSION),
    ]
    if audit.notes:
        audit_pairs.append(("Notas:", audit.notes))

    for label, value in audit_pairs:
        _write_info_pair(ws, row, label, value, col_start=1, col_end=5)
        row += 1

    row += 1

    # ── Resumen de hallazgos ────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="RESUMEN DE HALLAZGOS")
    apply_style(ws.cell(row=row, column=1), STYLE_SUBHEADER)
    row += 1

    sev_counts = {s: 0 for s in ["critical", "high", "medium", "low", "informational"]}
    for f in findings:
        if f.severity in sev_counts:
            sev_counts[f.severity] += 1

    headers = ["Crítico", "Alto", "Medio", "Bajo", "Informativo", "Total Hallazgos", "Dispositivos", "Estado"]
    _write_row(ws, row, headers, STYLE_HEADER)
    row += 1

    total_findings = sum(sev_counts.values())
    status = "CRÍTICO" if sev_counts["critical"] > 0 else ("RIESGO" if sev_counts["high"] > 0 else "REVISAR" if sev_counts["medium"] > 0 else "OK")
    values = [
        sev_counts["critical"], sev_counts["high"], sev_counts["medium"],
        sev_counts["low"], sev_counts["informational"],
        total_findings, len(devices), status,
    ]
    for col_idx, (val, sev) in enumerate(zip(values[:5], ["critical", "high", "medium", "low", "informational"]), 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        apply_style(cell, {**SEVERITY_STYLES[sev], "border": STYLE_HEADER["border"]})
    for col_idx, val in enumerate(values[5:], 6):
        cell = ws.cell(row=row, column=col_idx, value=val)
        apply_style(cell, STYLE_ROW_EVEN)

    # Ajuste columnas
    for i, w in enumerate([22, 30, 18, 14, 14, 18, 14, 12], 1):
        _set_col_width(ws, i, w)


def _sheet_inventario(wb, devices: list[Device]):
    ws = wb.create_sheet("Inventario")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = ["IP", "Hostname", "NetBIOS", "MAC", "Fabricante", "SO/Tipo", "Tipo Dispositivo", "Estado", "RTT (ms)", "Puertos Abiertos"]
    _write_row(ws, 1, headers, STYLE_HEADER)

    widths = [16, 28, 22, 18, 22, 22, 20, 10, 10, 14]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)

    for idx, d in enumerate(devices, 2):
        style = STYLE_ROW_EVEN if idx % 2 == 0 else STYLE_ROW_ODD
        values = [
            d.ip_address,
            d.hostname or "",
            d.netbios_name or "",
            d.mac_address or "",
            d.manufacturer or "",
            d.os_type or "",
            DEVICE_TYPE_LABELS.get(d.device_type, d.device_type),
            "Activo" if d.is_up else "Inactivo",
            round(d.response_time_ms, 1) if d.response_time_ms else "",
            len(d.ports),
        ]
        _write_row(ws, idx, values, style)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _sheet_red_puertos(wb, db: Session, devices: list[Device], audit_id: int):
    ws = wb.create_sheet("Red y Puertos")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = ["IP", "Hostname", "Puerto", "Protocolo", "Servicio", "Estado", "Banner", "Riesgo"]
    _write_row(ws, 1, headers, STYLE_HEADER)

    widths = [16, 28, 8, 10, 16, 10, 40, 10]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)

    row = 2
    for device in devices:
        ports = db.query(Port).filter(Port.device_id == device.id).order_by(Port.port_number).all()
        for port in ports:
            style = {**STYLE_ROW_EVEN} if row % 2 == 0 else {**STYLE_ROW_ODD}
            values = [
                device.ip_address, device.hostname or "",
                port.port_number, port.protocol.upper(),
                port.service or "", port.state.upper(),
                (port.banner or "")[:80],
                "⚠ RIESGO" if port.is_risky else "",
            ]
            _write_row(ws, row, values, style)
            if port.is_risky:
                from openpyxl.styles import PatternFill
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FFF3CD")
            row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _sheet_hallazgos(wb, findings: list[Finding], devices: list[Device]):
    ws = wb.create_sheet("Hallazgos")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    device_map = {d.id: d for d in devices}
    headers = ["Severidad", "Categoría", "Dispositivo", "Título", "Descripción", "Evidencia", "Recomendación", "Estado"]
    _write_row(ws, 1, headers, STYLE_HEADER)
    widths = [14, 16, 20, 45, 50, 40, 50, 14]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)

    order = {"critical": 0, "high": 1, "medium": 2, "low": 3, "informational": 4}
    sorted_findings = sorted(findings, key=lambda f: order.get(f.severity, 9))

    for idx, f in enumerate(sorted_findings, 2):
        device = device_map.get(f.device_id)
        device_label = device.ip_address if device else ""
        if device and device.hostname:
            device_label += f" ({device.hostname})"

        sev_style = SEVERITY_STYLES.get(f.severity, STYLE_ROW_EVEN)
        row_style = STYLE_ROW_EVEN if idx % 2 == 0 else STYLE_ROW_ODD

        values = [
            SEVERITY_LABELS.get(f.severity, f.severity.upper()),
            f.category.capitalize(),
            device_label,
            f.title,
            f.description or "",
            f.evidence or "",
            f.recommendation or "",
            f.status.capitalize(),
        ]
        _write_row(ws, idx, values, row_style)

        # Colorear columna de severidad
        sev_cell = ws.cell(row=idx, column=1)
        apply_style(sev_cell, {**sev_style, "border": STYLE_ROW_EVEN.get("border")})

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _sheet_recomendaciones(wb, findings: list[Finding]):
    ws = wb.create_sheet("Recomendaciones")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    ws.cell(row=1, column=1, value="RECOMENDACIONES Y PLAN DE ACCIÓN")
    apply_style(ws.cell(row=1, column=1), STYLE_SUBHEADER)

    headers = ["Prioridad", "Severidad", "Área", "Acción Recomendada", "Impacto", "Estado"]
    _write_row(ws, 2, headers, STYLE_HEADER)
    widths = [10, 14, 20, 70, 30, 14]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)

    priority_order = {"critical": 1, "high": 2, "medium": 3, "low": 4, "informational": 5}
    sorted_f = sorted(
        [f for f in findings if f.recommendation],
        key=lambda x: priority_order.get(x.severity, 9)
    )

    for idx, f in enumerate(sorted_f, 3):
        style = STYLE_ROW_EVEN if idx % 2 == 0 else STYLE_ROW_ODD
        values = [
            priority_order.get(f.severity, "-"),
            SEVERITY_LABELS.get(f.severity, ""),
            f.category.capitalize(),
            f.recommendation or "",
            "Alto" if f.severity in ("critical", "high") else "Medio",
            "Pendiente",
        ]
        _write_row(ws, idx, values, style)


def _sheet_comparativa(wb, comparison_data: dict):
    ws = wb.create_sheet("Comparativa")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    ws.cell(row=1, column=1, value="ANÁLISIS COMPARATIVO ENTRE REVISIONES")
    apply_style(ws.cell(row=1, column=1), STYLE_SUBHEADER)

    row = 3
    sections = [
        ("DISPOSITIVOS NUEVOS", comparison_data.get("new_devices", []), "new"),
        ("DISPOSITIVOS DESAPARECIDOS", comparison_data.get("removed_devices", []), "removed"),
        ("HALLAZGOS NUEVOS", comparison_data.get("new_findings", []), "finding_new"),
        ("HALLAZGOS RESUELTOS", comparison_data.get("resolved_findings", []), "finding_resolved"),
        ("HALLAZGOS PERSISTENTES", comparison_data.get("persistent_findings", []), "finding_persistent"),
    ]

    for section_title, items, section_type in sections:
        ws.merge_cells(f"A{row}:F{row}")
        ws.cell(row=row, column=1, value=f"{section_title} ({len(items)})")
        apply_style(ws.cell(row=row, column=1), STYLE_HEADER)
        row += 1

        if not items:
            ws.cell(row=row, column=1, value="Sin cambios detectados")
            apply_style(ws.cell(row=row, column=1), STYLE_VALUE)
            row += 2
            continue

        if "finding" in section_type:
            for item in items:
                style = STYLE_ROW_EVEN if row % 2 == 0 else STYLE_ROW_ODD
                ws.cell(row=row, column=1, value=item.get("severity", "").upper())
                ws.cell(row=row, column=2, value=item.get("title", ""))
                ws.cell(row=row, column=3, value=item.get("category", ""))
                for col in range(1, 4):
                    apply_style(ws.cell(row=row, column=col), style)
                row += 1
        else:
            for item in items:
                style = STYLE_ROW_EVEN if row % 2 == 0 else STYLE_ROW_ODD
                ws.cell(row=row, column=1, value=item.get("ip", ""))
                ws.cell(row=row, column=2, value=item.get("hostname", ""))
                ws.cell(row=row, column=3, value=item.get("device_type", ""))
                for col in range(1, 4):
                    apply_style(ws.cell(row=row, column=col), style)
                row += 1
        row += 1

    widths = [16, 40, 20, 20, 20, 20]
    for i, w in enumerate(widths, 1):
        _set_col_width(ws, i, w)


def _sheet_comparativa_empty(wb):
    ws = wb.create_sheet("Comparativa")
    ws.merge_cells("A1:D1")
    ws.cell(row=1, column=1, value="COMPARATIVA - Sin datos de revisión anterior")
    apply_style(ws.cell(row=1, column=1), STYLE_SUBHEADER)


def _sheet_placeholder(wb, name: str, msg: str = "Módulo disponible en próxima actualización"):
    ws = wb.create_sheet(name)
    ws.merge_cells("A1:D1")
    ws.cell(row=1, column=1, value=name.upper())
    apply_style(ws.cell(row=1, column=1), STYLE_SUBHEADER)
    ws.cell(row=3, column=1, value=msg)
    apply_style(ws.cell(row=3, column=1), STYLE_VALUE)


def _sheet_hardware(wb): _sheet_placeholder(wb, "Hardware")
def _sheet_backups(wb): _sheet_placeholder(wb, "Backups")
def _sheet_vmware(wb): _sheet_placeholder(wb, "VMware")
def _sheet_fortigate(wb): _sheet_placeholder(wb, "FortiGate")
def _sheet_snmp(wb): _sheet_placeholder(wb, "SNMP")
def _sheet_windows(wb): _sheet_placeholder(wb, "Windows")
def _sheet_nas(wb): _sheet_placeholder(wb, "NAS")
def _sheet_ilo_idrac(wb): _sheet_placeholder(wb, "iLO-iDRAC")


def _sheet_metadata(wb, audit, client, technician):
    """Hoja oculta con metadatos para importación entre instalaciones."""
    ws = wb.create_sheet("_auditcheck_meta")
    ws.sheet_state = "hidden"

    meta = [
        ("client_id", audit.client_id),
        ("client_name", client.name if client else ""),
        ("audit_id", audit.id),
        ("audit_date", audit.completed_at.isoformat() if audit.completed_at else datetime.utcnow().isoformat()),
        ("technician", technician.username if technician else ""),
        ("technician_full_name", technician.full_name if technician else ""),
        ("auditcheck_version", settings.APP_VERSION),
        ("exported_at", datetime.utcnow().isoformat()),
    ]

    ws.cell(row=1, column=1, value="key")
    ws.cell(row=1, column=2, value="value")
    for row_idx, (key, value) in enumerate(meta, 2):
        ws.cell(row=row_idx, column=1, value=key)
        ws.cell(row=row_idx, column=2, value=str(value))
